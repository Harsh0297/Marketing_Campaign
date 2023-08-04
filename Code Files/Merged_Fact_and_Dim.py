import pandas as pd 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook



file_name1="Data Files\Marketing_Campaign_Dimension.xlsx"
file_name2="Data Files\Marketing_Campaign_Fact.xlsx"

raw_df1=pd.read_excel(file_name1) 

raw_df2=pd.read_excel(file_name2)
# Dt_Customer is the time when the customer enrolled into the campaign
# It is in object type we have to convert them into Date Format 

raw_df1['Dt_Customer']=pd.to_datetime(raw_df1['Dt_Customer'],infer_datetime_format=False, dayfirst=True)
raw_df1['Enrolled_Date']=raw_df1['Dt_Customer'].dt.date
raw_df1=raw_df1.drop('Dt_Customer',axis=1)

raw_df1['Age']=2023-raw_df1['Year_Birth']
raw_df1=raw_df1.drop('Year_Birth',axis=1)

raw_df1.loc[(raw_df1['Marital_Status']=="Single"),"Marital_Status_People"] = "1"
raw_df1.loc[(raw_df1['Marital_Status']=="Together"),"Marital_Status_People"] = "2"
raw_df1.loc[(raw_df1['Marital_Status']=="Married"),"Marital_Status_People"] = "2"
raw_df1.loc[(raw_df1['Marital_Status']=="Divorced"),"Marital_Status_People"] = "1"
raw_df1.loc[(raw_df1['Marital_Status']=="Widow"),"Marital_Status_People"] = "1"
raw_df1.loc[(raw_df1['Marital_Status']=="Alone"),"Marital_Status_People"] = "1"
raw_df1.loc[(raw_df1['Marital_Status']=="Absurd"),"Marital_Status_People"] = "1"
raw_df1.loc[(raw_df1['Marital_Status']=="YOLO"),"Marital_Status_People"] = "1"






merge_file=pd.merge(raw_df1,raw_df2,on="ID",how="left")


merge_file.dropna(inplace=True)

wb=Workbook()
wb1=Workbook()

#Deleted predefined excel file 
del wb['Sheet']

# Created a worksheet name "Merged Data"
ws=wb.create_sheet("Merged_Fact_Dim")


# This is the bydefault function to write data into excel file 
for r in dataframe_to_rows(merge_file,index=False,header=True):
        ws.append(r)


# Save Excel file in XLSX format so tht i can upload it PowerBI for Visualization
wb.save("Data Files\Merged_Fact_Dim.xlsx")



