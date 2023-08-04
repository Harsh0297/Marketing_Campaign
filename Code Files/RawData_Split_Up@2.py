import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
file_name="Data Files\marketing_campaign.csv"

raw_df=pd.read_csv(file_name,sep="\t")

# print(raw_df.info())

raw_df2=raw_df[['ID','Year_Birth','Education','Marital_Status','Income','Kidhome','Teenhome','Dt_Customer','Recency']]

exluded_columns=raw_df[['Year_Birth','Education','Marital_Status','Income','Kidhome','Teenhome','Dt_Customer','Recency']]

raw_df3=raw_df.drop(columns=exluded_columns)

# print(raw_df2.info())

# print(raw_df3.info())



wb=Workbook()
wb1=Workbook()

#Deleted predefined excel file 
del wb['Sheet']
del wb1['Sheet']

# Created a worksheet name "Merged Data"
ws=wb.create_sheet("Marketing_Campaign_Dimension")
ws1=wb1.create_sheet("Marketing_Campaign_Fact")


# This is the bydefault function to write data into excel file 
for r in dataframe_to_rows(raw_df2,index=False,header=True):
        ws.append(r)

for r in dataframe_to_rows(raw_df3,index=False,header=True):
        ws1.append(r)

# Save Excel file in XLSX format so tht i can upload it PowerBI for Visualization
wb1.save("Data Files\Marketing_Campaign_Fact.xlsx")

wb.save("Data Files\Marketing_Campaign_Dimension.xlsx")

